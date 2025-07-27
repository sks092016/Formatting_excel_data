from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Load the existing Excel workbook

wb = load_workbook('/Users/subhashsoni/Documents/Bharatnet_OFC_planning/SHIVPURI-BADARWAS1.0/SHIVPURI-BADARWAS-1.0.xlsx')

# Define styles
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
font = Font(name='Cambria', size=11)
header_font = Font(bold=True, name='Cambria', size=12)

# Process each sheet
for ws in wb.worksheets:
    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Auto-adjust row heights (optional) and apply formatting
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = thin_border
            cell.font = header_font if row_idx == 1 else font
            if row_idx == 1:
                cell.fill = header_fill

# Save the formatted file
wb.save('formatted_output.xlsx')
