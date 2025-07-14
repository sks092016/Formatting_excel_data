import json
import pandas as pd
import geopandas as gpd
from math import radians, sin, cos, sqrt, atan2
import tkinter as tk
from tkinter import filedialog
import time
import difflib
import numpy as np
import shutil
from pathlib import Path
import re
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

root = tk.Tk()
root.withdraw()
with open('References/village_district_code.json', 'r') as f:
    village_data = json.load(f)
######################################### CONSTANTS #############################################################
# ROAD WIDTHS
PMGY = 4
SH = 15
NH = 40
GRAMPANCHAYAT = GP = 6
PWD = 6
ODR = 6
MDR = 6
NAGAR_PARISHAD = 4
MPRRDA = 4
OTHERS = 6

# RM & MH INTERVAL
rm_interval = 200
mh_interval = 1800

# PROTECTION
culvert_protection = 'DWC + PCC'
bridge_protection = 'GI + Clamping'
hard_rock = 'DWC + PCC'

# VERSION
version = '1.0'

# PATH
# for Windows
folder_path = 'D:\\bharat_net_data\\'
# for Mac
# folder_path = '/Users/subhashsoni/Documents/Bharatnet_OFC_planning/'

# APIKEY
api_key = 'AIzaSyBpsTQbW0ax0c18wGhC46wLkIPNvOH1sb4'

# DEFAULT LAT-LONG
lat = 0
lon = 0
place_id = ''
############################################### API URLS ##########################################################
# GOOGLE
place_url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={lat},{lon}&radius=50&key={api_key}"
roads_place_id_url = f"https://roads.googleapis.com/v1/nearestRoads?points={lat},{lon}&key={api_key}"
road_name_url = f"https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&fields=name&key={api_key}"

# OPENSM
url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lon}&format=json"
headers_village_name = {
    'User-Agent': 'village-lookup-script'
}
headers_road_name = {
    'User-Agent': 'road-name-fetcher-script'
}
################################################ To Be Used ########################################################

# pd.DataFrame([0.0] + [sum(list(temp_df['distance'])[:i + 1]) for i in range(1, len(list(temp_df['distance'])))]).values

######################################### Reference Files ########################################################

base_url = "https://fieldsurvey.rbt-ltd.com/app"
gdf_reference = gpd.read_file("References/Formats/BoQ & RoW.shp")

######################################## Reading the Shape File ##################################################

shapefile_path = filedialog.askopenfilename(
    title="Select a shape file",
    filetypes=[("Shape Files", "*.shp"), ("All files", "*.*")]
)
gdf_working = gpd.read_file(shapefile_path)

###################################### Creating the Directory for file store #####################################
blockName = gdf_working['block_name'][0]
districtName = gdf_working['district_n'][0]

dir_path = Path(folder_path + districtName + "-" + blockName + version)

# If it exists, delete it
if dir_path.exists() and dir_path.is_dir():
    shutil.rmtree(dir_path)  # Deletes the entire folder and its contents

# Now create it fresh
dir_path.mkdir(parents=True, exist_ok=False)

excel_file = str(dir_path) + f"\\{districtName}-{blockName}-{version}.xlsx"
###################################### Checking the Structure of the Files #######################################

is_structure_same = set(gdf_reference.columns) <= set(gdf_working.columns)
print(is_structure_same)
if not is_structure_same:
    col_diff_A = set(gdf_reference.columns) - set(gdf_working.columns)
    print("The Structure of the shape file is not matching the reference structure aborting. \n")
    time.sleep(0.5)
    print("Following columns are not present in the selected shape file:\n")
    time.sleep(0.5)
    print(col_diff_A)
else:
    print(gdf_working.columns)
###################################### Creating the Common Span Details #########################################

span_ = gdf_working[['from_gp_na', 'to_gp_name', 'span_name', 'ring_no', 'scope', 'span_id']].drop_duplicates()
gdf_working['distance'] = pd.to_numeric(gdf_working['distance'], errors='coerce')
span_dis = gdf_working.groupby('span_name').agg({'distance': 'sum'})
boq_sd_df = pd.merge(span_, span_dis, on=['span_name'], how='inner')

###################################### Creating the Details Sheet ################################################
def haversine_distance(lat1, lon1, lat2, lon2):
    """Calculate the distance between two lat-long points using the Haversine formula."""
    # Convert latitude and longitude to radians
    lat1, lon1, lat2, lon2 = map(radians, [float(lat1), float(lon1), float(lat2), float(lon2)])
    # Haversine formula
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    r = 6371e3  # Earth's radius in meters
    return c * r


def categorize_value(row):
    value = str(row['crossing_t'])  # Ensure value is string
    value2 = str(row['end_point_'])
    if any(sub in value.lower() for sub in ['road']):  # Check for 'Re' or 'RE'
        return 'Road'
    elif difflib.SequenceMatcher(None, value.lower(), "culvert").ratio() > 0.5:
        return 'Crossing'
    elif difflib.SequenceMatcher(None, value.lower(), "bridge").ratio() > 0.5:
        return 'Crossing'
    elif any(sub in value2.lower() for sub in ['gp', 'gram panchyat', 'grampanchayat']):  # Check for 'GP', 'Gp', etc.
        return 'Gram Panchyat'
    else:
        return 'Landmark'


def calculate_offset_width(row, param):
    try:
        value = str(row['road_autho']).upper()
        keywords = ['PMGY', 'SH', 'NH', 'NAGAR PARISHAD', 'GRAMPANCHAYAT', 'GP', 'PWD', 'ODR', 'MDR']
        matched = next((sub for sub in keywords if sub in value), None)
        if any(sub in value for sub in ['PMGY', 'SH', 'NH', 'Nagar Parishad', 'Grampanchyat', 'GP', 'PWD', 'ODR', 'MDR']):
            value = globals()[matched]
            if param == 'offset':
                return value / 2 + 0.2 * value
            if param == 'width':
                return value
        else:
            if param == 'offset':
                return OTHERS / 2 + 0.2 * OTHERS
            if param == 'width':
                return OTHERS
            return None
    except:
        return None


def finding_lat_lon(row, lat_lon):
    value_end_point = str(row['end_point_']).lower()
    value_crossing = str(row['crossing_t']).lower()
    att = lat_lon
    if att == 'lat':
        if any(sub in value_crossing for sub in ['culvert', 'bridge', 'road']):
            return row['Start_Lat']
        else:
            return row['End_Lat']
    elif att == 'lon':
        if any(sub in value_crossing for sub in ['culvert', 'bridge', 'road']):
            return row['Start_Long']
        else:
            return row['End_Long']
    return None


def calculating_rms(df):
    # Track cumulative distances
    cum_distance_start = 0
    rm_list = []
    for d in df['distance']:
        cum_distance_end = cum_distance_start + float(d)
        # Count new RMs in this segment
        rm_start_count = cum_distance_start // rm_interval
        rm_end_count = cum_distance_end // rm_interval
        new_rms = int(rm_end_count - rm_start_count)
        rm_list.append(new_rms)
        cum_distance_start = cum_distance_end
    rm_df = pd.DataFrame(rm_list, columns=['rm'], index=df.index)
    return rm_df


def calculating_mhs(df):
    cum_distance_start = 0
    mh_list = []
    for d in df['distance']:
        cum_distance_end = cum_distance_start + float(d)
        # Count new MHs in this segment
        mh_start_count = cum_distance_start // mh_interval
        mh_end_count = cum_distance_end // mh_interval
        new_mhs = int(mh_end_count - mh_start_count)
        mh_list.append(new_mhs)
        cum_distance_start = cum_distance_end
    mh_df = pd.DataFrame(mh_list, columns=['mh'], index=df.index)
    return mh_df


def calculate_road_chainage(row):
    chainage = ''
    if 'kms' in str(str(row['end_point_'])).lower():
        chainage = str(row['end_point_'])
    else:
        chainage = ''
    return chainage


def calculate_protec(row, param):
    value = str(row['crossing_t'])
    length = float(row['distance'])
    strata = str(row['strata_typ'])
    if difflib.SequenceMatcher(None, value.lower(), "culvert").ratio() > 0.5 and length <= 20:
        if param == 'struct' or param == 'for':
            return 'Culvert'
        elif param == 'type':
            return culvert_protection
        elif param == 'len':
            return length + 6
        elif param == 'length':
            return length
    elif difflib.SequenceMatcher(None, value.lower(), "bridge").ratio() > 0.5 and length > 20:
        if param == 'struct' or param == 'for':
            return 'Bridge'
        elif param == 'type':
            return bridge_protection
        elif param == 'len':
            return length + 6
        elif param == 'length':
            return length
    if strata == 'hard_rock' or strata == 'Hard Rock':
        if param == 'for':
            return 'Hard Rock'
        elif param == 'type':
            return hard_rock
        elif param == 'len':
            return length + 6
        elif param == 'length':
            return None
    return ''


def finding_utility(row, param):
    side = row['ofc_laying']
    value = str(row['end_point_']).lower()
    utility = None
    if 'rjil' in value:
        utility = "Reliance Jio"
    elif 'airtel' in value:
        utility = "Airtel"
    elif 'bsnl' in value:
        utility = 'BSNL'
    elif 'gas' in value:
        utility = 'Gas PipeLine ' + value
    elif 'hpcl' in value:
        utility = 'HPCL Pipeline'
    elif 'iocl' in value:
        utility = 'IOCL Pipeline'
    elif 'railway' in value:
        utility = 'railway'
    elif 'petrol' in value:
        utility = 'Petroleum Pipeline ' + value
    elif 'gail' in value:
        utility = 'Gail Xing'
    if param == 'utility':
        return utility
    if param == 'side' and utility is not None:
        return side
    return None

if is_structure_same:
    cols_ds = ['SPAN_CONTINUITY', 'POINT NAME', 'TYPE', 'POSITION', 'OFFSET', 'CHAINAGE', 'DISTENCE(M)', 'LATITUDE',
               "LONGITUDE", 'ROUTE NAME', 'ROUTE TYPE',
               'OFC TYPE', 'LAYING TYPE', 'ROUTE ID', 'ROUTE MARKER', 'MANHOLE', 'ROAD NAME', 'ROAD WIDTH(m)',
               'ROAD SURFACE', 'OFC POSITION', 'APRX DISTANCE FROM RCL(m)',
               'AUTHORITY NAME', 'ROAD CHAINAGE', 'ROAD STRUTURE TYPE', 'LENGTH (IN Mtr.)', 'PROTECTION TYPE',
               'PROTECTION FOR', 'PROTECTION LENGTH (IN Mtr.)', 'UTILITY NAME',
               'SIDE OF THE ROAD', 'SOIL TYPE', 'REMARKS']

    boq_ = pd.DataFrame(columns=cols_ds)
    span = gdf_working.sort_values('span_name').span_name.unique()
    for s in span:
        temp_df = gdf_working[gdf_working.span_name == s].sort_values('Sequqnce')
        boq_ds_df = pd.DataFrame(columns=cols_ds)
        boq_ds_df['SPAN_CONTINUITY'] = temp_df['Sequqnce']
        boq_ds_df['POINT NAME'] = temp_df['end_point_']
        if 'Type' in temp_df.columns:
            boq_ds_df['TYPE'] = temp_df['Type']
        else:
            boq_ds_df['TYPE'] = temp_df.apply(categorize_value, axis=1)
        boq_ds_df['POSITION'] = temp_df['ofc_laying']
        if 'ROW_Limit' in temp_df.columns:
            boq_ds_df['OFFSET'] = temp_df['ROW_Limit']
        else:
            boq_ds_df['OFFSET'] = temp_df.apply(calculate_offset_width, axis=1, args=('offset',))
        boq_ds_df['CHAINAGE'] = pd.DataFrame([sum(list(temp_df['distance'].astype(float))[:i - 1]) for i in range(1, len(list(temp_df['distance']))+1)]).values
        boq_ds_df['DISTENCE(M)'] = temp_df['distance']
        boq_ds_df['LATITUDE'] = temp_df.apply(finding_lat_lon, axis=1, args=('lat',))
        boq_ds_df['LONGITUDE'] = temp_df.apply(finding_lat_lon, axis=1, args=('lon',))
        boq_ds_df['ROUTE NAME'] = temp_df['span_name']
        boq_ds_df['ROUTE TYPE'] = temp_df['scope']
        boq_ds_df['OFC TYPE'] = '48F'
        boq_ds_df['LAYING TYPE'] = 'UG'
        boq_ds_df['ROUTE ID'] = temp_df['span_id']
        boq_ds_df['ROUTE MARKER'] = calculating_rms(temp_df)
        boq_ds_df['MANHOLE'] = calculating_mhs(temp_df)
        boq_ds_df['ROAD NAME'] = temp_df['road_name']
        if 'Road_Width' in temp_df.columns:
            boq_ds_df['ROAD WIDTH(m)'] = temp_df['Road_Width']
        else:
            boq_ds_df['ROAD WIDTH(m)'] = temp_df.apply(calculate_offset_width, axis=1, args=('width',))
        boq_ds_df['ROAD SURFACE'] = temp_df['road_surfa']
        boq_ds_df['OFC POSITION'] = temp_df['ofc_laying']
        if 'ROW_Limit' in temp_df.columns:
            boq_ds_df['APRX DISTANCE FROM RCL(m)'] = temp_df['ROW_Limit']
        else:
            boq_ds_df['APRX DISTANCE FROM RCL(m)'] = ''
        boq_ds_df['AUTHORITY NAME'] = temp_df['road_autho']
        boq_ds_df['ROAD CHAINAGE'] = temp_df.apply(calculate_road_chainage, axis=1)
        boq_ds_df['ROAD STRUTURE TYPE'] = temp_df.apply(calculate_protec, axis=1, args=('struct',))
        boq_ds_df['LENGTH (IN Mtr.)'] = temp_df.apply(calculate_protec, axis=1, args=('length',))
        boq_ds_df['PROTECTION TYPE'] = temp_df.apply(calculate_protec, axis=1, args=('type',))
        boq_ds_df['PROTECTION FOR'] = temp_df.apply(calculate_protec, axis=1, args=('for',))
        boq_ds_df['PROTECTION LENGTH (IN Mtr.)'] = temp_df.apply(calculate_protec, axis=1, args=('len',))
        boq_ds_df['UTILITY NAME'] = temp_df.apply(finding_utility, axis=1, args=('utility',))
        boq_ds_df['SIDE OF THE ROAD'] = temp_df.apply(finding_utility, axis=1, args=('side',))
        boq_ds_df['SOIL TYPE'] = temp_df['strata_typ']
        boq_ds_df['REMARKS'] = "NA"
        boq_ = pd.concat([boq_, boq_ds_df])

    with pd.ExcelWriter(excel_file, engine='openpyxl',
                        mode='w') as writer:
        boq_.to_excel(writer, sheet_name='Details Sheet', index=False)
    print('Details Sheet Created')
    # ############################################## Generating the Span Details ########################################

    cols_sd = ['FROM', 'TO', 'ROUTE NAME', 'RING NO.', 'ROUTE TYPE', 'OFC TYPE', 'LAYING TYPE', 'ROUTE ID',
               'TOTAL LENGTH(KM)',
               'OH', 'UG']
    span_details = pd.DataFrame(columns=cols_sd)

    span_details['FROM'] = boq_sd_df['from_gp_na']
    span_details['TO'] = boq_sd_df['to_gp_name']
    span_details['ROUTE NAME'] = boq_sd_df['span_name']
    span_details['RING NO.'] = boq_sd_df['ring_no']
    span_details['ROUTE TYPE'] = 'OFC to be laid for Ring Formation (in Km)'
    span_details['OFC TYPE'] = '48F'
    span_details['LAYING TYPE'] = 'UG'
    span_details['ROUTE ID'] = boq_sd_df['span_id']
    span_details['TOTAL LENGTH(KM)'] = boq_sd_df['distance'] / 1000
    span_details['OH'] = 0
    span_details['UG'] = boq_sd_df['distance'] / 1000

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',
                        if_sheet_exists='replace') as writer:
        span_details.to_excel(writer, sheet_name='Span Details', index=False)
    print('Span Details Created')
    # ############################################# Generating the RoW Details #########################################

    authorities = gdf_working['road_autho'].unique()

    cols_row = ['ROUTE NAME', 'ROUTE TYPE', 'RING NO', 'ROUTE ID', 'TOTAL ROUTE LENGTH'] + list(authorities)

    row_details = pd.DataFrame(columns=cols_row)

    df2 = gdf_working.pivot_table(values='distance', index='span_name', columns='road_autho', aggfunc='sum',
                                  fill_value=0).reset_index()
    row_ = pd.merge(boq_sd_df, df2, on=['span_name'], how='inner')

    row_details['ROUTE NAME'] = row_['span_name']
    row_details['ROUTE TYPE'] = row_['scope']
    row_details['RING NO'] = row_['ring_no']
    row_details['ROUTE ID'] = row_['span_id']
    row_details['TOTAL ROUTE LENGTH'] = row_['distance'] / 1000
    for auths in authorities:
        row_details[auths] = row_[auths] / 1000
    row_details['OTHERS'] = 0

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',
                        if_sheet_exists='replace') as writer:
        row_details.to_excel(writer, sheet_name='RoW', index=False)
    print('RoW Summery Created')
    ####################################### Generating the Protection Details ###########################################

    cols_pro = ['ROUTE NAME', 'ROUTE TYPE', 'TOTAL ROUTE LENGTH', 'NO OF CULVERT', 'LENGTH (IN Mtr) OF CULVERT',
                'NO OF BRIDGE',
                'LENGTH (IN Mtr) OF BRIDE',
                'LENGTH (IN Mtr) OF PCC', 'LENGTH (IN Mtr) OF DWC', 'LENGTH (IN Mtr) OF GI',
                'LENGTH (IN Mtr) OF DWC+PCC',
                "Soil Detail", "LENGTH (IN Mtr) OF DWC+PCC (HARD ROCK)", 'LENGTH (IN Mtr) OF ANCORING']
    protection_details = pd.DataFrame(columns=cols_pro)

    _details = pd.read_excel(str(dir_path) + f"/{districtName}-{blockName}-{version}.xlsx", sheet_name='Details Sheet')

    _protection = _details[['ROUTE NAME', 'ROAD STRUTURE TYPE', 'LENGTH (IN Mtr.)', 'PROTECTION TYPE', 'PROTECTION FOR',
                            'PROTECTION LENGTH (IN Mtr.)']]
    r_agg = _protection.groupby(['ROUTE NAME', 'PROTECTION FOR'])['PROTECTION LENGTH (IN Mtr.)'].agg(
        ['count', 'sum']).unstack(fill_value=0)
    if len(r_agg.columns) == 4:
        r_agg.columns = ['NO OF BRIDGE', 'NO OF CULVERT', 'LENGTH (IN Mtr) OF BRIDGE',
                         'LENGTH (IN Mtr) OF CULVERT']  # Rename columns
    else:
        r_agg.columns = ['NO OF BRIDGE', 'NO OF CULVERT', 'HARD ROCK', 'LENGTH (IN Mtr) OF BRIDGE',
                         'LENGTH (IN Mtr) OF CULVERT', 'HARD ROCK (Length)']  # Rename columns
    r_agg = r_agg.reset_index()
    p_agg = _protection.pivot_table(values='PROTECTION LENGTH (IN Mtr.)', index='ROUTE NAME', columns='PROTECTION TYPE',
                                    aggfunc='sum', fill_value=0).reset_index()
    protection_ = pd.merge(r_agg, p_agg, on='ROUTE NAME', how='outer')

    merged_df = boq_sd_df
    merged_df.rename(columns={'span_name': 'ROUTE NAME'}, inplace=True)
    protection_ = pd.merge(protection_, merged_df, on='ROUTE NAME', how='outer')

    protection_details['ROUTE NAME'] = protection_['ROUTE NAME']
    protection_details['ROUTE TYPE'] = 'OFC to be laid for Ring Formation (in Km)'
    protection_details['TOTAL ROUTE LENGTH'] = protection_['distance'] / 1000
    protection_details['NO OF CULVERT'] = protection_['NO OF CULVERT']
    protection_details['LENGTH (IN Mtr) OF CULVERT'] = protection_['LENGTH (IN Mtr) OF CULVERT'].astype(float) - 6 * protection_[
        'NO OF CULVERT'].astype(float)
    protection_details['NO OF BRIDGE'] = protection_['NO OF BRIDGE']
    protection_details['LENGTH (IN Mtr) OF BRIDE'] = protection_['LENGTH (IN Mtr) OF BRIDGE'].astype(float) - 6 * protection_[
        'NO OF BRIDGE'].astype(float)
    protection_details['LENGTH (IN Mtr) OF PCC'] = 0
    protection_details['LENGTH (IN Mtr) OF DWC'] = protection_['LENGTH (IN Mtr) OF CULVERT']
    protection_details['LENGTH (IN Mtr) OF GI'] = protection_['LENGTH (IN Mtr) OF BRIDGE']
    protection_details['LENGTH (IN Mtr) OF DWC+PCC'] = protection_['HARD ROCK (Length)']
    protection_details['Soil Details'] = ''
    protection_details['HARD ROCK (Length)'] = protection_['HARD ROCK (Length)']
    protection_details['LENGTH (IN Mtr) OF ANCORING'] = 0

    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a',
                        if_sheet_exists='replace') as writer:
        protection_details.to_excel(writer, sheet_name='Protection Details', index=False)

    print('Protection Summery Created')
    #
    ##################################################### RoW PreSurvey ######################################################
    #

    def extract_coords(geom):
        try:
            geom_str = str(geom)
            coords_part = geom_str.replace('LINESTRING (', '').replace(')', '')
            # Split by comma to separate coordinate pairs
            coords = [coord.strip() for coord in coords_part.split(',')]
            return coords
        except Exception as e:
            print(f"Error processing geom: {geom} -> {e}")
            return []


    cols_row_ps = ['SrNo', 'Ring No', 'GP Name', 'Span Name', 'NHSHNo', 'RoadWidth', 'RowBoundaryLmt', 'KMStoneFromA',
                   'KMStoneToB', 'SuveryDist', 'st_Lat_Long_Auth', 'end_Lat_Long_Auth', 'LandmarkRHS',
                   'VlgTwnPoint', 'OFClaying', 'RdCrossing Req', 'ReasonRdCrossing', 'UtilityLHS', 'UtilityChecked',
                   'RowAuthorityName', 'AuthorityAddress', 'FeasibilityOfROWApproval', 'TypeOfOverlapArea',
                   'NearestLandmark', 'LengthOfOverlapArea', 'ExpansionInProg', 'ExpansionPlanned', 'TypeOfCrossing',
                   'st_Lat_Long_xing', 'end_Lat_Long_xing', 'LatLandmark', 'LongLandmark',
                   'LengthOfCrossing', 'Remarks']

    span = gdf_working.sort_values('span_name').span_name.unique()
    row_pre_survey = pd.DataFrame(columns=cols_row_ps)
    for s in span:
        temp_df = gdf_working[gdf_working.span_name == s].sort_values('Sequqnce')
        output = []
        group = None
        crossing_coords_start = []
        crossing_coords_end = []
        output_df = None
        for _, row in temp_df.iterrows():
            p_n, s_n, dist, auth, geom, ofc_side, ring = row["crossing_t"], row["span_name"], row["distance"], row[
                "road_autho"], \
                row["geometry"], row['ofc_laying'], row['ring_no']
            coords = extract_coords(geom)
            if not any(sub in p_n.lower() for sub in ['culvert', 'bridge']):
                if group is None:
                    group = {
                        'SrNo': '',
                        "Ring No":ring,
                        'GP Name':'',
                        'Span Name': s_n,
                        'NHSHNo': '',
                        'RoadWidth': '',
                        'RowBoundaryLmt': '',
                        'KMStoneFromA': "",
                        'KMStoneToB': "",
                        'SuveryDist': dist,
                        'st_Lat_Long_Auth': f"{coords[0].split(' ')[1]},{coords[0].split(' ')[0]}",
                        'end_Lat_Long_Auth': f"{coords[-1].split(' ')[1]},{coords[-1].split(' ')[0]}",
                        'LandmarkRHS': "",
                        'VlgTwnPoint': "",
                        'OFClaying': ofc_side,
                        'RdCrossing Req': '',
                        'ReasonRdCrossing': '',
                        'UtilityLHS': '',
                        'UtilityRHS': "",
                        'UtilityChecked': '',
                        'RowAuthorityName': auth,
                        'AuthorityAddress': f"{blockName}, {districtName}",
                        'FeasibilityOfROWApproval': 'Yes',
                        'TypeOfOverlapArea': '',
                        'NearestLandmark': "",
                        'LengthOfOverlapArea': '',
                        'ExpansionInProg': '',
                        'ExpansionPlanned': '',
                        'TypeOfCrossing': '',
                        'st_Lat_Long_xing': '',
                        'end_Lat_Long_xing': '',
                        'LatLandmark': '',
                        'LongLandmark': '',
                        'LengthOfCrossing': '',
                        'Remarks': ''
                    }
                elif group["RowAuthorityName"] == auth:
                     group["SuveryDist"] += dist
                     group["end_Lat_Long_Auth"] = f"{coords[-1].split(' ')[1]},{coords[-1].split(' ')[0]}"
                else:
                    output.append(group)
                    group = {
                        'SrNo': '',
                        "Ring No": ring,
                        'GP Name': '',
                        'Span Name': s_n,
                        'NHSHNo': '',  # TODO Find Road Nmme
                        'RoadWidth': '',
                        'RowBoundaryLmt': '',
                        'KMStoneFromA': "",
                        'KMStoneToB': "",
                        'SuveryDist': dist,
                        'st_Lat_Long_Auth': f"{coords[0].split(' ')[1]},{coords[0].split(' ')[0]}",
                        'end_Lat_Long_Auth': f"{coords[-1].split(' ')[1]},{coords[-1].split(' ')[0]}",
                        'LandmarkRHS': "",  # TODO Find LandMark
                        'VlgTwnPoint': "",  # TODO Find Village from-to
                        'OFClaying': ofc_side,
                        'RdCrossing Req': '',
                        'ReasonRdCrossing': '',
                        'UtilityLHS': '',
                        'UtilityRHS': "",
                        'UtilityChecked': '',
                        'RowAuthorityName': auth,
                        'AuthorityAddress': f"{blockName}, {districtName}",
                        'FeasibilityOfROWApproval': 'Yes',
                        'TypeOfOverlapArea': '',
                        'NearestLandmark': "",  # TODO Find LandMark
                        'LengthOfOverlapArea': '',
                        'ExpansionInProg': '',
                        'ExpansionPlanned': '',
                        'TypeOfCrossing': '',
                        'st_Lat_Long_xing': '',
                        'end_Lat_Long_xing': '',
                        'LatLandmark': '',  # TODO Latitude of LandMark
                        'LongLandmark': '',  # TODO Longitude of the LandMark
                        'LengthOfCrossing': '',
                        'Remarks': ''
                    }
                    crossing_coords_start = []
                    crossing_coords_end = []
            else:
                if group:
                    group["Ring No"] = ring
                    group["TypeOfCrossing"] = p_n.lower()
                    crossing_coords_start.append(f"{coords[0].split(' ')[1]},{coords[0].split(' ')[0]}")
                    crossing_coords_end.append(f"{coords[-1].split(' ')[1]},{coords[-1].split(' ')[0]}")
                    group["SuveryDist"] += dist
                    group["end_Lat_Long_Auth"] = f"{coords[-1].split(' ')[1]},{coords[-1].split(' ')[0]}"
                    group["st_Lat_Long_xing"] = ", ".join(crossing_coords_start)
                    group["end_Lat_Long_xing"] = ", ".join(crossing_coords_end)
        # Append final group
        if group:
            output.append(group)
        # Create output DataFrame
        output_df = pd.DataFrame(output)
        output_df.reset_index(drop=True, inplace=True)
        row_pre_survey = pd.concat([row_pre_survey, output_df])


    def finding_road_name(row):
        print('*')
        lat, lon = row['st_Lat_Long_Auth'].split(',')[0],row['st_Lat_Long_Auth'].split(',')[1]
        roads_place_id_url = f"https://roads.googleapis.com/v1/nearestRoads?points={lat},{lon}&key={api_key}"
        response_pid = requests.get(roads_place_id_url)
        data = response_pid.json()
        road_name = []
        try:
            for loc in data['snappedPoints']:
                place_id = loc["placeId"]
                road_name_url = f"https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&fields=name&key={api_key}"
                response_pname = requests.get(road_name_url)
                data = response_pname.json()
                if data['result']['name'] == 'Unnamed Road':
                    road_name.append('')
                else:
                    road_name.append(data['result']['name'])
            return ", ".join(road_name)
        except:
            return road_name


    def finding_landmark(row, value):
        print('#')
        lat, lon = row['st_Lat_Long_Auth'].split(',')[0],row['st_Lat_Long_Auth'].split(',')[1]
        place_url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={lat},{lon}&radius=20&type=establishment&key={api_key}"
        response = requests.get(place_url)
        data = response.json()
        landmark = []
        geometry_lat = []
        geometry_lng = []
        try:
            for r in data['results']:
                landmark.append(r['name'])
                geometry_lat.append(str(r['geometry']['location']['lat']))
                geometry_lng.append(str(r['geometry']['location']['lng']))
            if value == 'name':
                return ', '.join(landmark)
            if value == 'lat':
                return ','.join(geometry_lat)
            if value == 'lng':
                return ','.join(geometry_lng)
        except:
            return None

    def finding_village(row):
        print('~')
        lat1, lon1 = row['st_Lat_Long_Auth'].split(',')[0],row['st_Lat_Long_Auth'].split(',')[1]
        place_url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={lat1},{lon1}&radius=20&key={api_key}"
        response_1 = requests.get(place_url)
        data_1 = response_1.json()
        lat2, lon2 = row['end_Lat_Long_Auth'].split(',')[0],row['end_Lat_Long_Auth'].split(',')[1]
        place_url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={lat2},{lon2}&radius=20&key={api_key}"
        response_2 = requests.get(place_url)
        data_2 = response_2.json()
        name = []
        try:
            for r in data_1['results']:
                name.append(r['name'])
            for r in data_2['results']:
                name.append(r['name'])
            return '-'.join(name).capitalize()
        except:
            return None
    row_pre_survey_temp = row_pre_survey
    if api_key is not None:
        row_pre_survey['RoadWidth'] = row_pre_survey_temp.apply(calculate_offset_width, axis=1, args=('width',))
        row_pre_survey['NHSHNo'] = row_pre_survey_temp.apply(finding_road_name, axis=1)
        row_pre_survey['LandmarkRHS'] = row_pre_survey_temp.apply(finding_landmark, axis=1, args=('name',))
        row_pre_survey['VlgTwnPoint'] = row_pre_survey_temp.apply(finding_village, axis=1)
        row_pre_survey['NearestLandmark'] = row_pre_survey_temp.apply(finding_landmark, axis=1, args=('name', ))
        row_pre_survey['LatLandmark'] = row_pre_survey_temp.apply(finding_landmark, axis=1, args=('lat',))
        row_pre_survey['LongLandmark'] = row_pre_survey_temp.apply(finding_landmark, axis=1, args=('lng',))
    else:
        print("GOOGLE API KEY is Required for generating all the details")
    with pd.ExcelWriter(excel_file, engine='openpyxl',
                        mode='a') as writer:
        row_pre_survey.to_excel(writer, sheet_name='PreSurvey', index=False)
    print('RoW Pre Survey Created')
########################################### Generating the Other Sheets ###############################################
############### INDEX ####################
data_ad = [
    {"Sl.No": 1, "Descrption": "ASSET DETAILS", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 2, "Descrption": "ANNEEXURE-X_TABLE A", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 3, "Descrption": "ANNEXURE X_TABLE-B", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 5, "Descrption": "SPAN DETAILS", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 6, "Descrption": "ROW", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 7, "Descrption": "LINE DIAGRAM", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 8, "Descrption": "Bill of Materials(BOM)", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 9, "Descrption": "Bill of Quantity(BOQ)", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 10, "Descrption": "Low Level Diagram(HLD)", "Status": "", "Remarks": ""},
    {"Sl.No": 11, "Descrption": "High Level Diagram(LLD)", "Status": "", "Remarks": ""},
    {"Sl.No": 12, "Descrption": "GPON", "Status": "Yes", "Remarks": ""},
    {"Sl.No": 13, "Descrption": "OTDR(Incremental)", "Status": "NA", "Remarks": ""},
    {"Sl.No": 14, "Descrption": "RM/MH Pictures(Incremental)", "Status": "NA", "Remarks": ""},
    {"Sl.No": 15, "Descrption": "BLOCK/GP Infra Pictures", "Status": "", "Remarks": ""},
    {"Sl.No": 16, "Descrption": "Route VideoGraphy", "Status": "", "Remarks": ""}
]

# Create DataFrame
df_index = pd.DataFrame(data_ad)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_index.to_excel(writer, sheet_name='Index', index=False)
print('Index Created')
################# ASSET DETAILS ####################
# Create survey detail data as dictionary
district_code = next((item["District Code"] for item in village_data["Report"] if item["District Name"] == districtName.capitalize()), None)
blockCode = next((item["Sub-District Code"] for item in village_data["Report"] if item["Sub-District Name"] == blockName.capitalize()), None)
total_length = boq_sd_df['distance'].sum() / 1000
# rings = boq_sd_df['ring_no'].unique()
data_asset = {
    "Field": [
        "BLOCK NAME",
        "BLOCK CODE",
        "DISTRICT NAME",
        "DISTRICT CODE",
        "STATE NAME",
        "STATE CODE",
        "Block Router",
        "GP Router",
        "NO OF RING",
        "TOTAL INCRIMENTAL CABLE TO BE USE",
        "TOTAL PROPOSED CABLE TO BE LAID",
        "Block Type"
    ],
    "Value": [
        blockName,
        blockCode,
        districtName,
        district_code,
        "Madhya Pradesh",
        "23",
        "1",
        '',
        '',
        "0.00 KM",
        f"{total_length} KM",
        "Green Field"
    ]
}

# Create DataFrame
df_asset = pd.DataFrame(data_asset)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_asset.to_excel(writer, sheet_name='Asset Details', index=False)
print('Index Created')
#################### Annexure -X
data_annex = {}
df_annex = pd.DataFrame(data_annex)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_annex.to_excel(writer, sheet_name='Annexure X', index=False)
print('Annexure X Created')
###################  Table B

# Define the column names exactly as per your table
columns_table_b = [
    "S.No.",
    "Block Name with LGD Code",
    "Name of GP",
    "GP LGD Code",
    "Availability of space Yes/ No",
    "Commercial Electric Supply Availability Yes/No",
    "Availability of Power Supply in Hrs.",
    "OLT Avalable/Not available",
    "Remarks",
    "Router"
]

# Create empty DataFrame with these columns
df_table_b = pd.DataFrame(columns=columns_table_b)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_table_b.to_excel(writer, sheet_name='Table B', index=False)
print('Table B Created')
###################  Gas Crossing
# Define the columns exactly as shown in your table
columns_gas = [
    "Sub Route Name",
    "Ring-Name",
    "Name of Gas/Oil pipeline Xing Co. (GAIL/BPCL/IOCL/IHB)",
    "Pipeline name as per signboard",
    "Gas/Oil pipeline Chainage as per signboard",
    "Latitude",
    "Longitude",
    "Landmark",
    "Village name",
    "SLD Status",
    "Tehsil name",
    "District name"
]

# Create empty DataFrame with these columns
df_gas = pd.DataFrame(columns=columns_gas)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_gas.to_excel(writer, sheet_name='Gas Crossing', index=False)
print('Gas Crossing Created')
###################### Railway Crossing
# Define column names exactly as in your table
columns_rlwy = [
    "Sr. No.",
    "Block/GP Link Name",
    "Sub Route Name",
    "Latitude",
    "Longitude",
    "Pole no. 1 LHS",
    "Pole no. 2 LHS",
    "Pole no.1 RHS",
    "Pole no.2 RHS",
    "Village name",
    "Tehsil name",
    "District name",
    "Length of total railway land involved in crossing (Mtrs)",
    "Distance of proposed crossing from electric pole (Mtrs)",
    "Height of Rails from formation level (Mtrs)",
    "Height of Rails from ground level (Mtrs)",
    "Name of both sides of railway station",
    "Total no. of railway tracks (Nos.)",
    "Pit Latlong (in)",
    "Pit Latlong (Out)"
]

# Create empty DataFrame with these columns
df_rlwy = pd.DataFrame(columns=columns_rlwy)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_rlwy.to_excel(writer, sheet_name='Railway Crossing', index=False)
print('Railway Crossing Created')
######################## GPON
columns_gpon = [
    "Block/GP Name",
    "Survey Location Type (Block/GP)",
    "Block/GP Address",
    "LGD Code",
    "Lat Long",
    "Location Type [BSNL Office/ BSNL Exchange/RSU/BTS/GP Office/School/ Other Building] in case of the other building pls specify the location type",
    "Location Secured and Locked/other. in case of the other pls specify",
    "Total Rack Available: 0/1/2",
    "FDF Installed [YES/NO]",
    "FDF Type [12F/24F/48F/other] in case of the other pls specify",
    "Terminated OFC Type 24/48F or other. in case of the other pls specify",
    "Nos of spare fiber available",
    "No of FDF PORT Used",
    "Make/Model",
    "Type",
    "Nos",
    "Types (SCM/PIC)",
    "Name",
    "Qty",
    "Working Status",
    "Equipment Type ((OLT/ONT))",
    "S.No",
    "Make",
    "Model",
    "CCU",
    "Battery",
    "OLT",
    "ONT",
    "Solar",
    "POWER Connection Status (Yes/No) if Yes please specify [1-Phase/3-Phase]",
    "Average Power Availability [No of Hours in a Day]",
    "POWER Back up Available (Yes/No)",
    "Solar/DC/AC Backup",
    "Capacity of backup in KVA/KWA",
    "Earth Pit-Lat/Long",
    "Strip (M) + Gauge",
    "Wire (M)",
    "ONT Installed / Commissioned [Yes/No]",
    "In case Yes, powered \"ON\" or \"OFF\"",
    "ZMH entry Location (Lat-Long)",
    "OFC entry Location (Lat-Long)"
]

# Create empty DataFrame with those columns
df_gpon = pd.DataFrame(columns=columns_gpon)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_gpon.to_excel(writer, sheet_name='GPON', index=False)

print('GPON Created')
############################ BoQ & BOM
bom = {}
df_bom = pd.DataFrame(bom)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_bom.to_excel(writer, sheet_name='BOM', index=False)

df_boq = pd.DataFrame(bom)
with pd.ExcelWriter(excel_file, engine='openpyxl',
                    mode='a') as writer:
    df_boq.to_excel(writer, sheet_name='BOQ', index=False)
print('BOM BOQ Created')
############################# Formatting the Excel Sheet
wb = load_workbook(excel_file)

# Define styles
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
font = Font(name='Cambria', size=11)
header_font = Font(bold=True, name='Cambria', size=12)

# Process each sheet
for ws in wb.worksheets:
    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Auto-adjust row heights (optional) and apply formatting
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell.border = thin_border
            cell.font = header_font if row_idx == 1 else font
            if row_idx == 1:
                cell.fill = header_fill

# Save the formatted file
wb.save(excel_file)
print('Process Completed')